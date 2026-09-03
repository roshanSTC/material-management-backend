from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties


CALCULATED_FIELDS = (
    "totalPriceEur", "totalPriceInr",
    "insuranceFreightEur", "insuranceFreightInr",
    "landedCostEur", "landedCostInr",
    "customsDutyInr", "igstInr", "transportationInr", "financingChargesInr",
    "totalCostInr", "lessIgstInr", "marginInr", "sellingPriceExclGst",
    "sellingPriceInclGst",
)

EXCEL_HEADERS = (
    "Pasaban Quotation Number", "Quotation Index", "Item Description", "Item Code",
    "Quantity", "Price Per Unit (EUR)", "Total Price (EUR)",
    "Insurance & Freight (EUR)", "Landed Cost (EUR)", "Landed Cost (INR)",
    "Customs & Duty (INR)", "IGST (INR)", "Transportation (INR)",
    "Financing Charges (INR)", "Total Cost (INR)", "Less IGST (INR)",
    "Margin (INR)", "Selling Price (excl. GST)", "Selling Price (incl. GST)",
)

GLOBAL_PARAMETER_ROWS = (
    ("EUR to INR", "eurToInr"),
    ("Insurance & Freight %", "insuranceFreightRate"),
    ("Default Customs Duty %", "defaultCustomsDutyRate"),
    ("IGST %", "igstRate"),
    ("Transportation %", "transportationRate"),
    ("Finance Charges %", "financeChargesRate"),
    ("Margin %", "marginRate"),
    ("GST %", "gstRate"),
)


def calculate_cost_sheet(*, global_params: dict, items: list[dict]) -> dict:
    """Calculate each cost component with production-grade Decimal precision."""
    calculated_items = [
        _calculate_item(global_params=global_params, item=item) for item in items
    ]
    column_totals = {
        field: float(sum(Decimal(str(item[field])) for item in calculated_items))
        for field in CALCULATED_FIELDS
    }
    total_selling_price_excl_gst = column_totals["sellingPriceExclGst"]
    total_gst = float(
        Decimal(str(total_selling_price_excl_gst)) * Decimal(str(global_params["gstRate"]))
    )
    grand_total_incl_gst = float(
        Decimal(str(total_selling_price_excl_gst)) + Decimal(str(total_gst))
    )
    return {
        "globalParams": global_params,
        "items": calculated_items,
        "columnTotals": column_totals,
        "totalSellingPriceExclGst": total_selling_price_excl_gst,
        "totalGst": total_gst,
        "grandTotalInclGst": grand_total_incl_gst,
    }


def build_cost_sheet_workbook(*, global_params: dict, items: list[dict]) -> BytesIO:
    """Build a workbook whose formulas recalculate when its inputs are edited."""
    workbook = Workbook()
    workbook.calculation = CalcProperties(
        calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True
    )
    worksheet = workbook.active
    worksheet.title = "Cost Sheet"
    worksheet.freeze_panes = "A11"
    _write_global_parameters(worksheet, global_params)
    _write_table(worksheet, items)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _calculate_item(*, global_params: dict, item: dict) -> dict:
    customs_duty_rate = item.get("customsDutyRate")
    if customs_duty_rate is None:
        customs_duty_rate = global_params["defaultCustomsDutyRate"]

    qty = Decimal(str(item["quantity"]))
    eur_to_inr = Decimal(str(global_params["eurToInr"]))
    ins_freight_rate = Decimal(str(global_params["insuranceFreightRate"]))
    cd_rate = Decimal(str(customs_duty_rate))
    igst_rate = Decimal(str(global_params["igstRate"]))
    trans_rate = Decimal(str(global_params["transportationRate"]))
    fin_rate = Decimal(str(global_params["financeChargesRate"]))
    margin_rate = Decimal(str(global_params["marginRate"]))
    gst_rate = Decimal(str(global_params["gstRate"]))

    # Resolve pricePerUnitEur and pricePerUnitInr
    if item.get("pricePerUnitEur") is not None:
        price_per_unit_eur = Decimal(str(item["pricePerUnitEur"]))
        price_per_unit_inr = price_per_unit_eur * eur_to_inr
    elif item.get("pricePerUnitInr") is not None:
        price_per_unit_inr = Decimal(str(item["pricePerUnitInr"]))
        price_per_unit_eur = price_per_unit_inr / eur_to_inr if eur_to_inr else price_per_unit_inr
    else:
        price_per_unit_eur = Decimal("0")
        price_per_unit_inr = Decimal("0")

    total_price_eur = price_per_unit_eur * qty
    total_price_inr = total_price_eur * eur_to_inr
    insurance_freight_eur = total_price_eur * ins_freight_rate
    insurance_freight_inr = insurance_freight_eur * eur_to_inr
    landed_cost_eur = total_price_eur + insurance_freight_eur
    landed_cost_inr = landed_cost_eur * eur_to_inr
    customs_duty_inr = landed_cost_inr * cd_rate
    igst_inr = (landed_cost_inr + customs_duty_inr) * igst_rate
    transportation_inr = (
        landed_cost_inr + customs_duty_inr + igst_inr
    ) * trans_rate
    financing_charges_inr = (
        landed_cost_inr + customs_duty_inr + igst_inr + transportation_inr
    ) * fin_rate
    total_cost_inr = (
        landed_cost_inr + customs_duty_inr + igst_inr + transportation_inr
        + financing_charges_inr
    )
    less_igst_inr = total_cost_inr - igst_inr
    margin_inr = less_igst_inr * margin_rate
    selling_price_excl_gst = less_igst_inr + margin_inr
    selling_price_incl_gst = selling_price_excl_gst * (Decimal("1") + gst_rate)

    return {
        **item,
        "pricePerUnitEur": float(price_per_unit_eur),
        "pricePerUnitInr": float(price_per_unit_inr),
        "totalPriceEur": float(total_price_eur),
        "totalPriceInr": float(total_price_inr),
        "insuranceFreightEur": float(insurance_freight_eur),
        "insuranceFreightInr": float(insurance_freight_inr),
        "landedCostEur": float(landed_cost_eur),
        "landedCostInr": float(landed_cost_inr),
        "customsDutyInr": float(customs_duty_inr),
        "igstInr": float(igst_inr),
        "transportationInr": float(transportation_inr),
        "financingChargesInr": float(financing_charges_inr),
        "totalCostInr": float(total_cost_inr),
        "lessIgstInr": float(less_igst_inr),
        "marginInr": float(margin_inr),
        "sellingPriceExclGst": float(selling_price_excl_gst),
        "sellingPriceInclGst": float(selling_price_incl_gst),
    }


def _write_global_parameters(worksheet, global_params: dict) -> None:
    label_fill = PatternFill("solid", fgColor="1F4E78")
    value_fill = PatternFill("solid", fgColor="D9EAF7")
    for row, (label, key) in enumerate(GLOBAL_PARAMETER_ROWS, start=1):
        label_cell = worksheet.cell(row=row, column=1, value=label)
        value_cell = worksheet.cell(row=row, column=2, value=global_params[key])
        label_cell.font = Font(color="FFFFFF", bold=True)
        label_cell.fill = label_fill
        value_cell.fill = value_fill
        value_cell.number_format = "#,##0.00" if row == 1 else "0.00%"
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 20


def _write_table(worksheet, items: list[dict]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for column, header in enumerate(EXCEL_HEADERS, start=1):
        cell = worksheet.cell(row=10, column=column, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    worksheet.cell(row=10, column=20, value="Customs Duty Rate Override")
    worksheet.column_dimensions["T"].hidden = True
    for row, item in enumerate(items, start=11):
        _write_item_row(worksheet, row, item)
    last_item_row = 10 + len(items)
    _write_totals(worksheet, first_item_row=11, last_item_row=last_item_row)
    worksheet.auto_filter.ref = f"A10:S{last_item_row}"
    _set_column_widths(worksheet)


def _write_item_row(worksheet, row: int, item: dict) -> None:
    for column, key in enumerate(
        ("quotationNumber", "quotationIndex", "itemDescription", "itemCode"), start=1
    ):
        worksheet.cell(row=row, column=column, value=item[key].strip())
    worksheet.cell(row=row, column=5, value=item["quantity"])
    worksheet.cell(row=row, column=6, value=item["pricePerUnitEur"])
    if item.get("customsDutyRate") is not None:
        rate_cell = worksheet.cell(
            row=row,
            column=20,
            value=item["customsDutyRate"],
        )
        rate_cell.number_format = "0.00%"
    formulas = {
        7: f"=F{row}*E{row}", 8: f"=G{row}*B$2", 9: f"=G{row}+H{row}",
        10: f"=I{row}*B$1", 11: _customs_duty_formula(row, item),
        12: f"=(J{row}+K{row})*B$4", 13: f"=(J{row}+K{row}+L{row})*B$5",
        14: f"=SUM(J{row}:M{row})*B$6", 15: f"=SUM(J{row}:N{row})",
        16: f"=O{row}-L{row}", 17: f"=P{row}*B$7", 18: f"=P{row}+Q{row}",
        19: f"=R{row}*(1+B$8)",
    }
    for column, formula in formulas.items():
        worksheet.cell(row=row, column=column, value=formula)
    worksheet.cell(row=row, column=5).number_format = "#,##0.000"
    for column in range(6, 10):
        worksheet.cell(row=row, column=column).number_format = "€#,##0.00"
    for column in range(10, 20):
        worksheet.cell(row=row, column=column).number_format = "₹#,##0.00"


def _customs_duty_formula(row: int, item: dict) -> str:
    customs_duty_rate = item.get("customsDutyRate")
    if customs_duty_rate is None:
        return f"=J{row}*B$3"
    return f"=J{row}*T{row}"


def _write_totals(worksheet, *, first_item_row: int, last_item_row: int) -> None:
    total_row = last_item_row + 1
    worksheet.cell(row=total_row, column=1, value="Column Totals").font = Font(bold=True)
    for column in range(7, 20):
        letter = get_column_letter(column)
        cell = worksheet.cell(
            row=total_row, column=column,
            value=f"=SUM({letter}{first_item_row}:{letter}{last_item_row})",
        )
        cell.font = Font(bold=True)
        cell.number_format = "€#,##0.00" if column < 10 else "₹#,##0.00"
    selling_total_row, gst_total_row, grand_total_row = total_row + 2, total_row + 3, total_row + 4
    summaries = (
        (selling_total_row, "Total Selling Price (excl. GST)", f"=SUM(R{first_item_row}:R{last_item_row})"),
        (gst_total_row, "Total GST", f"=B{selling_total_row}*B$8"),
        (grand_total_row, "Grand Total (incl. GST)", f"=B{selling_total_row}+B{gst_total_row}"),
    )
    for row, label, formula in summaries:
        worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        value_cell = worksheet.cell(row=row, column=2, value=formula)
        value_cell.font = Font(bold=True)
        value_cell.number_format = "₹#,##0.00"


def _set_column_widths(worksheet) -> None:
    widths = {"A": 26, "B": 18, "C": 36, "D": 18, "E": 12, "F": 18}
    widths.update({get_column_letter(column): 22 for column in range(7, 20)})
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
